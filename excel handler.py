import openpyxl
import statistics as st
import math
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active

row = sheet.max_row
column = sheet.max_column
l = []
for j in range(2, 176):
    l = []
    for i in range(2, 682):
        cell_obj = sheet.cell(row=j, column=i)
        if cell_obj.value == sheet['ZH' + str(j)].value:
            l.append(sheet.cell(row=1, column=i).value)

    sheet['ZG' + str(j)].value = math.ceil(st.median(l))

wb.save(filename="E:\\Materials Project\\DeepLearningModle\\test.xlsx")